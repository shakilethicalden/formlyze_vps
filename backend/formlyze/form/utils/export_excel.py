import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from io import BytesIO

def generate_excel(form, responses):

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Responses"
    
    # Get all unique question titles from responses
    question_titles = set()
    for response in responses:
        for item in response.response_data:
            if 'question_title' in item:
                question_titles.add(item['question_title'])
    

    question_titles = sorted(list(question_titles))
    

    headers = ['Response ID', 'Responder Email', 'Submission Date'] + question_titles
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header
        ws[f'{col_letter}1'].font = Font(bold=True)
    

    for row_num, response in enumerate(responses, 2):

        ws[f'A{row_num}'] = response.id
        ws[f'B{row_num}'] = response.responder_email or ''
        ws[f'C{row_num}'] = response.created_at.strftime('%Y-%m-%d %H:%M:%S')
        
        response_map = {}
        for item in response.response_data:
            if 'question_title' in item:
                response_value = item['response']
                
        
                if isinstance(response_value, dict):
                    response_value = ', '.join(f"{k}: {v}" for k, v in response_value.items())
                elif isinstance(response_value, list):
                    response_value = ', '.join(response_value)
                
                response_map[item['question_title']] = str(response_value)
        

        for col_num, question_title in enumerate(question_titles, 4):
            col_letter = get_column_letter(col_num)
            answer = response_map.get(question_title, '')
            ws[f'{col_letter}{row_num}'] = answer
    
    # adjust column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    
    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)
    virtual_workbook.seek(0)
    
    return virtual_workbook.getvalue()